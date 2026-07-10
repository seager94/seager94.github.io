#!/usr/bin/env python3
"""
export_lessonmap_json.py — export lessonmap.xlsx (master) to lessonmap.json

Purpose
-------
1. Produces the machine-readable export the curriculum index page needs
   (PROJECT_LOG build-backlog item 5).
2. Doubles as a structural validator: --check runs integrity checks on the
   master map and exits non-zero on hard failures, so it can gate a merge.

Usage (from repo root, master map on main)
------------------------------------------
    python export_lessonmap_json.py                 # writes lessonmap.json
    python export_lessonmap_json.py --check         # validate only, no write
    python export_lessonmap_json.py --map path.xlsx --out site/lessonmap.json

Conventions honoured
--------------------
- Loads read-only with data_only=True (values, not formulas) — the master is
  never written by this script, so the Excel-lock silent-save trap can't bite.
- Filename is taken from the Filename column when present, otherwise derived
  from the Published URL basename (19 rows currently rely on this).
- Strand-code drift (spa vs space, prb vs pro) is REPORTED, not failed —
  the standing rule is match-the-siblings until the whole-repo reconciliation.
"""

import argparse
import json
import re
import sys
from collections import Counter
from datetime import date, datetime

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("openpyxl is required:  pip install openpyxl")

EXPECTED_HEADERS = [
    "Lesson ID", "Year", "Strand", "Lesson #", "Topic (snake_case)",
    "Filename", "AC9 Descriptor(s)", "Elaboration #s", "Mapping note",
    "SA Conceptual Understanding", "Lesson focus", "Practice style",
    "Prerequisites", "Status", "Generated date", "Published URL", "Notes",
]

URL_RE = re.compile(r"^https://seager94\.github\.io/(year-\d+|methods-stage-[12])/[a-z\-]+/[\w\-.]+\.html$")
FNAME_RE = re.compile(r"^(y(7|8|9|10)_[a-z]+_\d{2}_[a-z0-9_]+(_v\d+)?|methods_s[12]_\d{2}_[a-z0-9_]+(_v\d+)?)\.html$")


def cell(row, idx, key):
    v = row[idx[key]] if idx.get(key) is not None and idx[key] < len(row) else None
    if isinstance(v, (datetime, date)):
        return v.isoformat()[:10]
    return str(v).strip() if v is not None else ""


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--map", default="lessonmap.xlsx")
    ap.add_argument("--out", default="lessonmap.json")
    ap.add_argument("--check", action="store_true", help="validate only; do not write JSON")
    args = ap.parse_args()

    wb = load_workbook(args.map, read_only=True, data_only=True)
    if "Lessons" not in wb.sheetnames:
        sys.exit(f"FAIL: no 'Lessons' sheet in {args.map}")
    ws = wb["Lessons"]

    hdr, idx = None, {}
    lessons, errors, warnings = [], [], []
    seen_ids = set()

    for row in ws.iter_rows(values_only=True):
        if hdr is None:
            hdr = [str(h).strip() if h else "" for h in row]
            idx = {h: i for i, h in enumerate(hdr)}
            missing = [h for h in EXPECTED_HEADERS if h not in idx]
            if missing:
                sys.exit(f"FAIL: Lessons sheet missing expected columns: {missing}")
            continue
        if not row or not row[0]:
            continue

        lid = cell(row, idx, "Lesson ID")
        status = cell(row, idx, "Status")
        url = cell(row, idx, "Published URL")
        fname = cell(row, idx, "Filename") or (url.rsplit("/", 1)[-1] if url else "")

        rec = {
            "id": lid,
            "year": cell(row, idx, "Year"),
            "strand": cell(row, idx, "Strand"),
            "lesson_num": cell(row, idx, "Lesson #"),
            "topic": cell(row, idx, "Topic (snake_case)"),
            "filename": fname,
            "ac9_descriptors": [d.strip() for d in re.split(r"[;,]\s*", cell(row, idx, "AC9 Descriptor(s)")) if d.strip()],
            "elaborations": cell(row, idx, "Elaboration #s"),
            "mapping_note": cell(row, idx, "Mapping note"),
            "sa_conceptual_understanding": cell(row, idx, "SA Conceptual Understanding"),
            "lesson_focus": cell(row, idx, "Lesson focus"),
            "practice_style": cell(row, idx, "Practice style"),
            "prerequisites": [p.strip() for p in cell(row, idx, "Prerequisites").split(",") if p.strip()],
            "status": status,
            "generated_date": cell(row, idx, "Generated date"),
            "url": url,
        }
        lessons.append(rec)

        # ---- integrity checks -------------------------------------------
        if lid in seen_ids:
            errors.append(f"duplicate Lesson ID: {lid}")
        seen_ids.add(lid)

        if status == "Published":
            if not url:
                errors.append(f"{lid}: Published but no URL")
            elif not URL_RE.match(url):
                warnings.append(f"{lid}: URL off-pattern: {url}")
            if "TODO" in rec["elaborations"]:
                warnings.append(f"{lid}: Published with TODO elaborations (known debt)")
            if fname and not FNAME_RE.match(fname):
                warnings.append(f"{lid}: filename off-convention: {fname}")
        elif status == "Planned":
            if url:
                errors.append(f"{lid}: Planned but has URL — status/URL disagree")
        elif status:
            warnings.append(f"{lid}: unrecognised status '{status}'")

        # prerequisite references must exist (checked after full read)

    id_set = {l["id"] for l in lessons}
    for l in lessons:
        for p in l["prerequisites"]:
            if p and p not in id_set:
                warnings.append(f"{l['id']}: prerequisite '{p}' not a known Lesson ID")

    tallies = Counter(l["status"] for l in lessons)
    by_year_strand = Counter((l["year"], l["strand"], l["status"]) for l in lessons)

    # ---- report ---------------------------------------------------------
    print(f"Rows: {len(lessons)}   " + "   ".join(f"{k}: {v}" for k, v in sorted(tallies.items())))
    if errors:
        print(f"\nERRORS ({len(errors)}):")
        for e in errors:
            print(f"  ✗ {e}")
    if warnings:
        print(f"\nWarnings ({len(warnings)}):")
        for w in warnings:
            print(f"  ! {w}")
    if not errors and not warnings:
        print("Integrity: clean.")

    if not args.check:
        out = {
            "generated": datetime.now().isoformat(timespec="seconds"),
            "source": args.map,
            "tallies": dict(tallies),
            "coverage": [
                {"year": y, "strand": s, "status": st, "count": c}
                for (y, s, st), c in sorted(by_year_strand.items())
            ],
            "lessons": lessons,
        }
        with open(args.out, "w", encoding="utf-8") as f:
            json.dump(out, f, ensure_ascii=False, indent=1)
        print(f"\nWrote {args.out} ({len(lessons)} lessons).")

    sys.exit(1 if errors else 0)


if __name__ == "__main__":
    main()
