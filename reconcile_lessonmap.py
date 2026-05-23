#!/usr/bin/env python3
"""
reconcile_lessonmap.py

Rebuilds the 'Lessons' sheet of the lesson map from the ACTUAL published
lesson HTML files in the repo. The source of truth is the <meta> tags that
the interactive-html-maths-lesson skill embeds in every lesson.

Why this exists: the hand-maintained 'Status' column drifted out of sync
with reality - Year 7 lessons marked 'Planned' that are actually live, and
Probability rows using the wrong id prefix. A map generated straight from
the repo cannot drift.

Run it from the repo root:   python reconcile_lessonmap.py
or pass a path:              python reconcile_lessonmap.py C:\\path\\to\\repo

It writes a NEW file (never overwrites your master map) and prints a
discrepancy report so you can see exactly what changed.
"""

import os
import re
import sys
import glob
from collections import Counter

import openpyxl

# ------------------------------ CONFIG --------------------------------
REPO_ROOT  = "."                          # repo root (or pass as argv[1])
OLD_MAP    = "lessonmap.xlsx"           # your current map - set the real name/path
NEW_MAP    = "lessonmap_reconciled.xlsx"  # output - reviewed before replacing master
PAGES_BASE = "https://seager94.github.io"
EXCLUDE_DIRS = {"lesson-drafts", ".claude", ".git", "node_modules"}
# Columns the HTML cannot supply - carried over from OLD_MAP, matched by Lesson ID
CARRY_OVER = ["Practice style", "Prerequisites", "Generated date", "Notes"]
# -----------------------------------------------------------------------

HEADERS = ['Lesson ID', 'Year', 'Strand', 'Lesson #', 'Topic (snake_case)',
           'Filename', 'AC9 Descriptor(s)', 'Elaboration #s', 'Mapping note',
           'SA Conceptual Understanding', 'Lesson focus', 'Practice style',
           'Prerequisites', 'Status', 'Generated date', 'Published URL', 'Notes']

# map column                       <- meta tag name
META_MAP = {
    'Lesson ID'                   : 'lesson-id',
    'Year'                        : 'sa-year',
    'Strand'                      : 'sa-strand',
    'AC9 Descriptor(s)'           : 'ac9-descriptor',
    'Elaboration #s'              : 'ac9-elaborations',
    'Mapping note'                : 'mapping-note',
    'SA Conceptual Understanding' : 'sa-conceptual-understanding',
    'Lesson focus'                : 'lesson-focus',
}

STRAND_ORDER = ['Number', 'Algebra', 'Measurement', 'Space',
                'Statistics', 'Probability']


def read_meta(html):
    """Pull <meta name=.. content=..> tags from a lesson file."""
    meta = {}
    for m in re.finditer(
            r'<meta\s+name=["\']([^"\']+)["\']\s+content=["\'](.*?)["\']\s*/?>',
            html, re.IGNORECASE | re.DOTALL):
        meta[m.group(1).strip()] = m.group(2).strip()
    return meta


def topic_from_filename(fname):
    """y7_sta_03_stem_and_leaf_plots_v2.html -> stem_and_leaf_plots"""
    stem = re.sub(r'\.html?$', '', fname, flags=re.IGNORECASE)
    stem = re.sub(r'_v\d+$', '', stem, flags=re.IGNORECASE)
    parts = stem.split('_')
    if len(parts) > 3 and re.match(r'^y\d+$', parts[0], re.IGNORECASE):
        parts = parts[3:]
    return '_'.join(parts)


def lesson_number(lesson_id, fname):
    """Two-digit lesson number from the id or filename."""
    for src in (lesson_id or '', fname or ''):
        m = re.search(r'_(\d{2})(?:_|$)', src)
        if m:
            return m.group(1)
    return ''


def year_sort(y):
    """Numeric years sort numerically; Methods S1/S2 etc. sort after."""
    try:
        return (0, int(y))
    except (ValueError, TypeError):
        return (1, str(y))


def carry_over_lookup(path):
    """{lesson_id: {col: value}} for hand-entered columns in the old map."""
    lookup = {}
    if not os.path.isfile(path):
        print(f"  note: old map '{path}' not found - "
              f"carry-over columns will be left blank.")
        return lookup
    wb = openpyxl.load_workbook(path, data_only=True)
    if 'Lessons' not in wb.sheetnames:
        return lookup
    ws = wb['Lessons']
    hdr = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(hdr)}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[idx.get('Lesson ID', 0)] in (None, ''):
            continue
        lid = str(row[idx['Lesson ID']]).strip()
        lookup[lid] = {c: row[idx[c]] for c in CARRY_OVER if c in idx}
    return lookup


def main():
    repo = sys.argv[1] if len(sys.argv) > 1 else REPO_ROOT
    print(f"Scanning repo for lesson HTML files under: {os.path.abspath(repo)}")

    lessons, skipped, incomplete = [], [], []
    for path in sorted(glob.glob(os.path.join(repo, '**', '*.html'),
                                 recursive=True)):
        rel = os.path.relpath(path, repo).replace(os.sep, '/')
        if rel.split('/')[0] in EXCLUDE_DIRS:
            continue
        with open(path, encoding='utf-8', errors='replace') as fh:
            meta = read_meta(fh.read())
        if 'lesson-id' not in meta:
            skipped.append(rel)                 # not a tagged lesson
            continue
        fname = os.path.basename(path)
        row = {h: '' for h in HEADERS}
        for col, tag in META_MAP.items():
            row[col] = meta.get(tag, '')
        row['Filename']           = fname
        row['Topic (snake_case)'] = topic_from_filename(fname)
        row['Lesson #']           = lesson_number(meta.get('lesson-id'), fname)
        row['Status']             = 'Published'
        row['Published URL']      = f"{PAGES_BASE}/{rel}"
        missing = [t for t in META_MAP.values() if not meta.get(t)]
        if missing:
            incomplete.append((rel, missing))
        lessons.append(row)

    if not lessons:
        print("\nNo tagged lesson HTML found - is the repo path correct?")
        sys.exit(1)

    # carry over hand-entered columns from the old map
    carry = carry_over_lookup(OLD_MAP)
    for row in lessons:
        for col, val in carry.get(str(row['Lesson ID']).strip(), {}).items():
            row[col] = '' if val is None else val

    # sort: year, strand, lesson number
    def sort_key(r):
        try:
            s = STRAND_ORDER.index(r['Strand'])
        except ValueError:
            s = 99
        return (year_sort(r['Year']), s, str(r['Lesson #']))
    lessons.sort(key=sort_key)

    # write new workbook - keep the other sheets, replace 'Lessons'
    if os.path.isfile(OLD_MAP):
        wb = openpyxl.load_workbook(OLD_MAP)
        if 'Lessons' in wb.sheetnames:
            del wb['Lessons']
        ws = wb.create_sheet('Lessons', 1)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Lessons'
    ws.append(HEADERS)
    for row in lessons:
        ws.append([row[h] for h in HEADERS])
    wb.save(NEW_MAP)

    # ----------------------------- report ------------------------------
    print(f"\nReconciled map written: {NEW_MAP}")
    print(f"Lessons found in repo : {len(lessons)}\n")
    for (yr, st), n in sorted(Counter((str(r['Year']), r['Strand'])
                                      for r in lessons).items(),
                              key=lambda kv: (year_sort(kv[0][0]),
                                              kv[0][1])):
        print(f"  Year {yr:<4} {st:<13} {n}")

    if skipped:
        print(f"\nHTML files with no lesson-id meta (ignored): {len(skipped)}")
        for p in skipped:
            print(f"  - {p}")
    if incomplete:
        print(f"\nLessons MISSING meta tags ({len(incomplete)}) - check these:")
        for rel, miss in incomplete:
            print(f"  - {rel}: missing {', '.join(miss)}")

    if carry:
        old_ids = set(carry)
        new_ids = {str(r['Lesson ID']).strip() for r in lessons}
        only_repo = sorted(new_ids - old_ids)
        only_map  = sorted(old_ids - new_ids)
        if only_repo:
            print(f"\nIn repo but NOT in old map ({len(only_repo)}) - new/renamed:")
            for i in only_repo:
                print(f"  + {i}")
        if only_map:
            print(f"\nIn old map but NOT in repo ({len(only_map)}) - "
                  f"planned, renamed, or deleted:")
            for i in only_map:
                print(f"  - {i}")

    print(f"\nDone. Review {NEW_MAP}, then replace your master map if it looks right.")


if __name__ == '__main__':
    main()
